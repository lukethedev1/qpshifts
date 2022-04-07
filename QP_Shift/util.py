from functools import partial
from django import forms
from django.conf import settings
from datetime import timedelta
from dateutil import tz
from django.core.files.temp import NamedTemporaryFile
from django.http import HttpResponse
from django.template.loader import get_template
from openpyxl import Workbook
from io import BytesIO
from xhtml2pdf import pisa

DateInput = partial(forms.DateInput, {'class': 'datepicker'})
DateTimeInput = partial(forms.DateTimeInput, {'class': 'datetimepicker', 'autocomplete': 'off'})


def render_to_pdf(template_src, context_dict=None):
    if context_dict is None:
        context_dict = {}

    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()

    pdf = pisa.pisaDocument(BytesIO(html.encode('ISO-8859-1')), dest=result)

    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')

    return None


def generar_excel_detalle(records):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Record Details'

    headers = [
        'Name',
        'Email',
        'Phone',
        'Check in',
        'Check out',
    ]

    row_num = 1

    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for record in records:

        correo = ''

        if hasattr(record.user_location.user, 'usuario'):
            nombre = record.user_location.user.get_full_name()
            correo = record.user_location.user.email
            telefono = record.user_location.user.usuario.telefono

        fecha_ingreso = "{0:%d-%m-%Y %H:%M}".format(obtener_datetime_sin_timezone(record.date_from)) if record.date_from else ''
        fecha_salida = "{0:%d-%m-%Y %H:%M}".format(obtener_datetime_sin_timezone(record.date_until)) if record.date_until else ''

        row_num += 1

        row = [
            nombre,
            correo,
            telefono,
            fecha_ingreso,
            fecha_salida
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    with NamedTemporaryFile() as temporary_file:

        workbook.save(temporary_file.name)
        temporary_file.seek(0)
        response = HttpResponse(temporary_file.read(), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=records.xlsx'

    return response


def generar_excel(records, filter_form):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Records'

    headers = [
        'Name',
        'Email',
        'Phone',
        'Locations',
        'Worked Hours',
    ]

    row_num = 1

    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for record in records:

        correo = ''

        if hasattr(record.user_location.user, 'usuario'):
            nombre = record.user_location.user.get_full_name()
            correo = record.user_location.user.email
            telefono = record.user_location.user.usuario.telefono

        locations = ""

        for location in record.user_location.user.locations.all():
            locations += location.location.name
            if not location.location.id == record.user_location.user.locations.all().last().id:
                locations += ","

        worked_hours = record.user_location.user.usuario.worked_hours(filter_form, record.user_location.user)

        row_num += 1

        row = [
            nombre,
            correo,
            telefono,
            locations,
            worked_hours
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    with NamedTemporaryFile() as temporary_file:

        workbook.save(temporary_file.name)
        temporary_file.seek(0)
        response = HttpResponse(temporary_file.read(), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=records.xlsx'

    return response


def obtener_datetime_sin_timezone(fecha):
    nueva_fecha = fecha.replace(tzinfo=tz.gettz(settings.TIME_ZONE))
    return nueva_fecha + timedelta(seconds=nueva_fecha.utcoffset().total_seconds())
